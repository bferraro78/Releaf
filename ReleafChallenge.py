# Releaf Coding Challenge - Benjamin Ferraro
# Python 2.7
import pyexcel
import feedparser
from openpyxl import load_workbook
import re # regex
import string

## Set up companyToRow dictionary = {'compnay' -> rowNum in SAMPLE_DATA}
my_array = pyexcel.get_array(file_name="rev_rest_africa.xlsx")
companyToRow = {}
count = 1
for row in my_array:
	if row[0] != "" and row[0] != "Name":
		companyToRow[row[0]] = count
	count+=1

# Set up URL dictionary {'company' - > [URL1, URL2,...]}
URLs = {}
for company in companyToRow:
	URLs[company] = []

# 141 categories from allAfrica.com
###### SHRINK THIS ARRAY IF TIMEOUT ERROR IS OCCURING
RSSCategories = ["latest", "africa", "centralafrica", "eastafrica", "northafrica", "southernafrica", "westafrica", "algeria", "angola", "benin", "botswana", "burkinafaso", "burundi", "cameroon", "capeverde", "centralafricanrepublic", "chad", "comoros", "congo_brazzaville", "congo_kinshasa", "cotedivoire", "djibouti", "egypt", "equatorialguinea", "eritrea", "ethiopia", "gabon", "gambia", "ghana", "guinea", "guineabissau", "kenya", "lesotho", "liberia", "libya", "madagascar", "malawi", "mali", "mauritania", "mauritius", "morocco", "mozambique", "namibia", "niger", "nigeria", "rwanda", "senegal", "seychelles", "sierraleone", "somalia", "southafrica", "southsudan", "sudan", "swaziland", "saotomeandprincipe", "tanzania", "togo", "tunisia", "uganda", "westernsahara", "zambia", "zimbabwe", "agoa", "onthemove", "agribusiness", "aid", "armsandarmies", "asiaaustrailiaandafrica", "arts", "athletics", "banking", "bookreviews", "books", "capitalflows", "children", "climate", "commodities", "company", "conflict", "construction", "corruption", "currencies", "debt", "ebola", "business", "ecotourism", "education", "environment", "energy", "europeandafrica", "externalrelations", "agriculture", "gameparks", "governance", "aids", "health", "humanrights", "ict", "infrastructure", "innovation", "io", "terroism", "investment", "labour", "land", "latinamericaandafrica", "legalaffairs", "malaria", "manufacturing", "middleeastandafrica", "migration", "mining", "music", "musicreviews", "nepad", "ngo", "nutrition", "oceans", "olympics", "peacekeeping", "petroleum", "polio", "pregnancy", "media", "privatization", "refugees", "religion", "science", "soccer", "sport", "stockmarkets", "sustainable", "trade", "transport", "travel", "tuberculosis", "usafrica", "urbanissues", "water", "wildlife", "women", "worldcup"]

numOfArticals = 0 # NUMBER OF ARTICALS COUNTER

allLinks = [] # Holds all links in all RSS feeds
# Parse RSS Feeds by category
for cat in RSSCategories:
	feed = "http://allafrica.com/tools/headlines/rdf/"+cat+"/headlines.rdf"
	d = feedparser.parse(feed)
	print(feed)

	# Grabs each entry's link inside an RSS Feed for a whole category
	for e in d.entries:
		# Increment artical counter
		numOfArticals += 1
		allLinks.append(e.link)

# Parses all links
for link in allLinks:
	dt = feedparser.parse(link)
	print(link)
	# Sanitize HTML page by removing tags, puncuation, whitespace, and newlines
	summary = dt.feed.summary
	summary = re.sub('<[^>]*>', ' ', summary).replace('\n', '')
	summary = re.sub('[^A-Za-z0-9]+', ' ', summary)

	# If any of the company names in 
	# SAMPLE_DATA exist in the articals word
	# array, add it to the company's URL dictionary.
	for company in companyToRow:
		if company in summary:
			if link not in URLs[company]:
				URLs[company].append(link)

print("Number of Articals Parsed: " + str(numOfArticals))

## START UPLOADING URLS TO DATA FILE
wb = load_workbook("rev_rest_africa.xlsx")
ws = wb.active

ws['L1'] = "RELAVANT URL" # ADD URL COLUMN

## For each company, grab the array of URLs
# and put it in the company's row cell
for company in URLs:
	col = 12
	row = companyToRow[company]
	if len(URLs[company]) != 0:
		for l in URLs[company]:
			cell = ws.cell(row=row, column=col)
			cell.value = l
			col+=1
	else:
		cell = ws.cell(row=row, column=col)
		cell.value = ""

# Save the excel file
wb.save('rev_rest_africa.xlsx')